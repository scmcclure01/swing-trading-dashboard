"""
Swing Trading Weekly Playbook Template — v4.0
Formatting library for weekly playbook Excel files.
Framework v4.0: Core-Satellite Structure + Velocity Flag + Deployment Floor

EXPECTED TAB STRUCTURE (v4.0)
─────────────────────────────
1. DASHBOARD         — Summary: regime, permission, Core status, deployment floor, actions
2. L0 - Macro Regime — Full Layer 0 readings including Velocity Flag (ROC 21 per sector)
3. L2 - Permission   — Permission state, week's rules, Core allocation by state
4. L3 - Sector Rot — Sector rotation ETF status, flow readings, Phase 1/2 tracking
5. CORE ALLOCATION   — [v4] Core ETF positions, % deployed, vs floor target, entry/exit status
6. L4 - Selection    — Stock screening: full signal, half signal, sector priority
7. L5 - Entry        — Entry triggers: standard, pullback, AND Accelerating Protocol [v4]
8. L6 - Sizing       — Position sizing for Tactical, Core, and Accelerating entries
9. L7 - Exposure     — Portfolio heat, deployment floor check [v4], concentration limits
10. L8 - Management  — Open position management: stops, targets, Core trail, Accelerating holds
11. L9 - Risk Caps   — Drawdown analysis with recalibrated tiers [v4]

CORE ALLOCATION TAB CONTENT [v4]
────────────────────────────────
- Current Core positions (ETF, entry, stop at 20d MA, % of account)
- Core target vs actual (GREEN=40%, YELLOW=20%, RED=0%)
- Deployment floor status (total deployed vs minimum)
- Phase 2 sectors eligible for new Core entry
- Core exit signals (20d MA, RS, regime, flows)

VELOCITY FLAG SECTION (in L0 tab) [v4]
──────────────────────────────────────
- ROC 21 for each sector ETF
- Flag status: ACCELERATING (>15%) / NORMAL (5-15%) / SLOW (<5%)
- Accelerating sectors trigger modified Layer 5 entry rules

FORMATTING GUIDELINES
─────────────────────
1. Gridlines: ALWAYS disable gridlines on every sheet.
   Call: disable_gridlines(ws) after creating each sheet.

2. Row heights & column widths: ALWAYS verify text fits.
   - Long text cells (Action, Notes, Reason columns): minimum row height 30.0; use 36–45 for multi-line.
   - Short label rows (headers, tickers, numbers): 15.75–19.5 is fine.
   - If a value string exceeds ~40 chars in a narrow column, widen the column or increase the row height.
   - Wrap text is ON by default for all data cells — row height must be tall enough to show wrapped lines.
   - Rule of thumb: 1 line ≈ 15pt height; 2 lines ≈ 30pt; 3 lines ≈ 45pt.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Color palette ─────────────────────────────────────────────────────────────
C_DARK_BLUE  = 'FF1D3B93'
C_LIGHT_BLUE = 'FF93B6FA'
C_YELLOW     = 'FFFFFECE'
C_GREEN      = 'FFE2EFDA'
C_PURPLE     = 'FF7030A0'
C_WHITE      = 'FFFFFFFF'
C_BLACK      = 'FF000000'

# ── Style primitives ──────────────────────────────────────────────────────────
def _font(name='Arial', size=10, bold=False, color=C_BLACK):
    return Font(name=name, size=size, bold=bold, color=color)

def _fill(rgb):
    return PatternFill('solid', fgColor=rgb)

def _align(horiz='left', vert='center', wrap=False):
    return Alignment(horizontal=horiz, vertical=vert, wrap_text=wrap)

# ── Cell writers ──────────────────────────────────────────────────────────────
def write(ws, coord, val, font=None, fill=None, align=None):
    cell = ws[coord]
    cell.value = val
    if font:  cell.font      = font
    if fill:  cell.fill      = fill
    if align: cell.alignment = align
    return cell

# ── Named styles (matching archive exactly) ───────────────────────────────────

def title(ws, coord, val):
    """Row 2 title — Arial 14 bold white / dark-blue, centered"""
    write(ws, coord, val,
          font=_font(size=14, bold=True, color=C_WHITE),
          fill=_fill(C_DARK_BLUE),
          align=_align('center'))

def subtitle(ws, coord, val):
    """Row 3 subtitle — Arial 13 bold white / dark-blue, centered"""
    write(ws, coord, val,
          font=_font(size=13, bold=True, color=C_WHITE),
          fill=_fill(C_DARK_BLUE),
          align=_align('center'))

def subtitle_light(ws, coord, val):
    """Row 3 light variant — Arial 12 bold dark-blue / light-blue, centered"""
    write(ws, coord, val,
          font=_font(size=12, bold=True, color=C_DARK_BLUE),
          fill=_fill(C_LIGHT_BLUE),
          align=_align('center'))

def section_hdr(ws, coord, val, bg=C_DARK_BLUE):
    """Section banner — Arial 10 bold white / dark-blue (or purple), left"""
    write(ws, coord, val,
          font=_font(size=10, bold=True, color=C_WHITE),
          fill=_fill(bg),
          align=_align('left'))

def col_hdr(ws, coord, val, bg=C_LIGHT_BLUE, fc=C_DARK_BLUE):
    """Column header — Arial 9 bold dark-blue / light-blue, centered, wrap"""
    write(ws, coord, val,
          font=_font(size=9, bold=True, color=fc),
          fill=_fill(bg),
          align=_align('center', wrap=True))

def col_hdr_purple(ws, coord, val):
    """Column header on purple — Arial 9 bold white / purple, centered, wrap"""
    col_hdr(ws, coord, val, bg=C_PURPLE, fc=C_WHITE)

def data(ws, coord, val, size=9, bold=False, bg=C_WHITE, fc=C_BLACK,
         horiz='left', wrap=True, num_fmt=None):
    """Standard data cell — Arial 9 (or 10), black on white, left, wrap"""
    cell = write(ws, coord, val,
                 font=_font(size=size, bold=bold, color=fc),
                 fill=_fill(bg),
                 align=_align(horiz, wrap=wrap))
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def data_center(ws, coord, val, size=9, bold=False, bg=C_WHITE, fc=C_BLACK, num_fmt=None):
    return data(ws, coord, val, size=size, bold=bold, bg=bg, fc=fc,
                horiz='center', wrap=True, num_fmt=num_fmt)

def data_yellow(ws, coord, val, size=9, bold=False, horiz='left'):
    return data(ws, coord, val, size=size, bold=bold, bg=C_YELLOW,
                horiz=horiz, wrap=True)

def data_yellow_center(ws, coord, val, size=9, bold=False):
    return data_yellow(ws, coord, val, size=size, bold=bold, horiz='center')

def data_green(ws, coord, val, size=9, bold=False, horiz='left'):
    return data(ws, coord, val, size=size, bold=bold, bg=C_GREEN,
                horiz=horiz, wrap=True)

def spacer_fill(ws, row, max_col=11):
    """Fill spacer row cells with dark-blue to match archive visual"""
    for col in range(1, max_col+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(C_WHITE)

# ── Sheet structural helpers ──────────────────────────────────────────────────
def disable_gridlines(ws):
    """Turn off gridlines for a sheet. Call this immediately after creating every sheet."""
    ws.sheet_view.showGridLines = False

def set_col_widths(ws, widths_dict):
    """widths_dict: {'A': 1.0, 'B': 18.5, ...}
    GUIDELINE: verify longest cell value fits the column width.
    Increase width if text > ~40 chars in columns that don't wrap (e.g. ticker, price columns).
    For wide text columns (Notes, Action, Reason) prefer width >= 30 and wrap_text=True with adequate row height.
    """
    for col, w in widths_dict.items():
        ws.column_dimensions[col].width = w

def set_row_heights(ws, heights_dict):
    """heights_dict: {2: 31.5, 3: 27.75, ...}
    GUIDELINE: match height to expected line count for wrapped cells.
      1 line  → 15.75 – 18.0
      2 lines → 27.75 – 30.0
      3 lines → 40.0  – 45.0
    Always increase rather than clip — never let text overflow hidden below the row boundary.
    """
    for row, h in heights_dict.items():
        ws.row_dimensions[row].height = h

def merge(ws, rng):
    ws.merge_cells(rng)

# ── Convenience: write a full data row across specified columns ───────────────
def data_row(ws, row, col_vals, bg=C_WHITE, size=9, bold=False):
    """col_vals: list of (col_letter, value) tuples"""
    for col, val in col_vals:
        data(ws, f'{col}{row}', val, size=size, bold=bold, bg=bg)

def data_row_center(ws, row, col_vals, bg=C_WHITE, size=9, bold=False):
    for col, val in col_vals:
        data_center(ws, f'{col}{row}', val, size=size, bold=bold, bg=bg)

